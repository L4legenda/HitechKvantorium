
from openpyxl import load_workbook, drawing, Workbook, worksheet
import qrcode

wb_form = load_workbook("data/data2.xlsx")

# seria = input("Серия: ")
# counts = int(input("Количество: "))
# num = int(input("Номер: ").strip())
# money = input("Цена: ")
""" 
qr = qrcode.QRCode(
    version=1,
    error_correction=qrcode.constants.ERROR_CORRECT_L,
    box_size=7,
    border=1,
)
qr.add_data(input("QR code: "))
qr.make(fit=True)

img = qr.make_image(fill_color="black", back_color="white")
img.save("qr/qrcode.png")
"""
list1 = wb_form["Лист1"]
cols = "ABCDEFGHIJKLMNO"
wb = Workbook()
ws = wb_form.active


elements = {
    "A1" : list1["A1"],
    "B1" : list1["B1"],
    "C1" : list1["C1"],
    "D1" : list1["D1"],
    "D5" : list1["D5"],
    "D9" : list1["D9"],
    "D13" : list1["D13"],
    "E1" : list1["E1"],
    "E9" : list1["E9"],
    "G1" : list1["G1"],
    "G2" : list1["G2"],
    "G3" : list1["G3"],
    "G6" : list1["G6"],
    "G8" : list1["G8"],
    "G10" : list1["G10"],
    "F13" : list1["F13"],
    "I6" : list1["I6"],
    "I8" : list1["I8"],
    "I10" : list1["I10"],
    "I15" : list1["I15"],
    "K1" : list1["K1"],
    "L11" : list1["L11"],
    "P1" : list1["P1"]
}

merge_cells = [
    ["A1", "A16"],
    ["B1", "B16"],
    ["C1", "C16"],
    ["D1", "D4"],
    ["D5", "D8"],
    ["D9", "D12"],
    ["D13", "D16"],
    ["E1", "E8"],
    ["E9", "E16"],
    ["G1", "I1"],
    ["G2", "I2"],
    ["G3", "I4"],
    ["G6", "H7"],
    ["G8", "H9"],
    ["G10", "H11"],
    ["I6", "I7"],
    ["I8", "I9"],
    ["I10", "I11"],
    ["F13", "I13"],
    ["I15", "N16"],
    ["K1", "O1"],
    ["P1", "P16"],
    ["L11", "N11"],
    ["K2", "O10"],
]


def render_sheet(ind = 1):

    for i in range(1, ind + 1):

        for j in merge_cells:
            p1 = j[0][0] + str(int(j[0][1:]) + (16 * i))
            p2 = j[1][0] + str(int(j[1][1:]) + (16 * i))
            ws.merge_cells(p1 + ":" + p2)

        for j in elements:
            id = j[0] + str( int(j[1:]) + (16 * i) )
            list1[id] = elements[j].value
            from_style = list1[j]._style
            list1[id]._style = from_style
            list1[id].hyperlink = list1[j].hyperlink
            list1[id].number_format = list1[j].number_format

        for j in range(1, 10):
            ws.row_dimensions[j + (16 * i)].height = 25




render_sheet(4)
print("Good")


#my_png = drawing.image.Image("qr/qrcode.png")
#ws.add_image(my_png, "K3")


wb_form.save("export.xlsx")