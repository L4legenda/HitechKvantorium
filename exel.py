
from openpyxl import load_workbook, drawing, Workbook, worksheet
import qrcode




def render_sheet(counts, seria, num, money, iqrcode):
    wb_form = load_workbook("data/data2.xlsx")

    counts = int(counts)
    num = int(num.strip())

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=8,
        border=1,
    )
    qr.add_data(iqrcode)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    img.save("qr/qrcode.png")

    list1 = wb_form["Лист1"]
    wb = Workbook()
    ws = wb_form.active

    list1["D1"] = num
    list1["D9"] = seria
    list1["E1"] = money + " ₽"

    list1["I6"] = seria
    list1["I8"] = num
    list1["I10"] = money + " ₽"

    my_png = drawing.image.Image("qr/qrcode.png")
    ws.add_image(my_png, "K" + str(2))

    elements = {
        "A1": list1["A1"],
        "B1": list1["B1"],
        "C1": list1["C1"],
        "D1": list1["D1"],
        "D5": list1["D5"],
        "D9": list1["D9"],
        "D13": list1["D13"],
        "E1": list1["E1"],
        "E9": list1["E9"],
        "G1": list1["G1"],
        "G2": list1["G2"],
        "G3": list1["G3"],
        "G6": list1["G6"],
        "G8": list1["G8"],
        "G10": list1["G10"],
        "F13": list1["F13"],
        "I6": list1["I6"],
        "I8": list1["I8"],
        "I10": list1["I10"],
        "I15": list1["I15"],
        "K1": list1["K1"],
        "L11": list1["L11"],
        "P1": list1["P1"]
    }

    merge_cells = [
        ["A1", "A16"],
        ["B1", "B16"],
        ["C1", "C16"],
        ["D1", "D4"],
        ["D5", "D8"],
        ["D9", "D12"],
        ["D13", "D16"],
        ["E1", "E8"],
        ["E9", "E16"],
        ["G1", "I1"],
        ["G2", "I2"],
        ["G3", "I4"],
        ["G6", "H7"],
        ["G8", "H9"],
        ["G10", "H11"],
        ["I6", "I7"],
        ["I8", "I9"],
        ["I10", "I11"],
        ["F13", "I14"],
        ["I15", "N16"],
        ["K1", "O1"],
        ["P1", "P16"],
        ["L11", "N11"],
        ["K2", "O10"],
    ]

    step = 0
    for i in range(1, counts + 1):
        if (i) % 4 == 0:
            step += 3
        for j in elements:
            id = j[0] + str( int(j[1:]) + (16 * i) + step )
            list1[id] = elements[j].value
            list1["D" + str((i * 16) + 1 + step) ] = num + i
            list1["I" + str((i * 16) + 8 + step)] = num + i
            from_style = list1[j]._style
            list1[id]._style = from_style
            list1[id].hyperlink = list1[j].hyperlink
            list1[id].number_format = list1[j].number_format

        list1["F" + str((16 * i) + 1 + step)]._style = list1["F1"]._style
        list1["F" + str((16 * i) + 16 + step)]._style = list1["F16"]._style
        list1["G" + str((16 * i) + 16 + step)]._style = list1["G16"]._style
        list1["H" + str((16 * i) + 16 + step)]._style = list1["H16"]._style

        my_png = drawing.image.Image("qr/qrcode.png")
        ws.add_image(my_png, "K" + str(2 + (16 * i) + step))

        for j in merge_cells:
            p1 = j[0][0] + str(int(j[0][1:]) + (16 * i) + step)
            p2 = j[1][0] + str(int(j[1][1:]) + (16 * i) + step)
            ws.merge_cells(p1 + ":" + p2)
        for j in range(1, 17):
            ws.row_dimensions[j + (16 * i) + step].height = ws.row_dimensions[j].height
            list1["P" + str((16 * i) + j + step)]._style = list1["P" + str(j)]._style

    print("Good")

    list1.print_area = "A1:P" + str((16 * counts) + step + 16)

    wb_form.save("export.xlsx")